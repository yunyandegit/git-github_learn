import openpyxl             #install openpyxl using code: pip install openpyxl
from datetime import date

wb= openpyxl.load_workbook('SE_class.xlsx')     #load existing excel file 

sheet=wb.active
#open the active sheet

today=date.today()

fh=open("chat.txt","r+")                #open file to read the input data i.e. the names of students from chats
data=fh.readlines()

for line in data:

    line=line.strip()
    print(line)                         #print input line
    nrp=line.split('|')                 #break it into name, roll no. , present word

    try:                                # code to check if the given roll no. is a NUMBER
        roll=int(nrp[1].strip())
    except:
        continue

    if(len(nrp)==3 and nrp[2].strip()=="Present"):

        n=sheet.max_row+1           #number of rows in the excel sheet
        # print(n)
        for i in range(3,n):
            if(sheet.cell(row=i,column=1).value==nrp[0].strip() and int(sheet.cell(row=i,column=2).value)==int(nrp[1].strip())):
                print(True)
                #print True if there matches a student details with the input one's

                sheet.cell(row=i, column=today.day+1).value="Present"
                #update the excel sheet data
                break

            #save at last the updated excel sheet
wb.save('SE_class.xlsx')
